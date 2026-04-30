import logging

from django.contrib.auth import get_user_model
from django.db import transaction
from openpyxl import load_workbook

from accounts.services import generate_temp_password
from imports.models import ImportJob, ImportRowLog
from notifications.services import send_credentials_email
from records.services import create_or_update_record

logger = logging.getLogger('imports')
User = get_user_model()
REQUIRED_COLUMNS = ['full_name', 'email', 'phone', 'address', 'dob']


COLUMN_ALIASES = {
    'full_name': ['full_name', 'fullname', 'full name', 'name', 'employee name', 'employee_name', 'emp_name', 'emp name'],
    'email': ['email', 'email address', 'email_address', 'emailaddress', 'e-mail', 'mail'],
    'phone': ['phone', 'phone number', 'phone_number', 'phonenumber', 'mobile', 'mobile number', 'mobile_number', 'contact', 'contact number', 'tel', 'telephone'],
    'address': ['address', 'home address', 'home_address', 'street address', 'street_address', 'location', 'addr'],
    'dob': ['dob', 'date of birth', 'date_of_birth', 'dateofbirth', 'birth date', 'birth_date', 'birthdate', 'birthday'],
}


def _normalize_headers(raw_headers):
    return [str(h).strip().lower() if h else '' for h in raw_headers]


def _map_headers(headers):
    mapped = list(headers)
    for canonical, aliases in COLUMN_ALIASES.items():
        for i, h in enumerate(headers):
            if h in aliases:
                mapped[i] = canonical
                break
    return mapped


def _row_to_dict(headers, row_values):
    data = {}
    for idx, header in enumerate(headers):
        data[header] = row_values[idx] if idx < len(row_values) else None
    return data


@transaction.atomic
def process_import_file(file_obj, uploaded_by):
    logger.info('[IMPORT] Starting import job for file "%s" by %s', file_obj.name, uploaded_by)
    job = ImportJob.objects.create(file_name=file_obj.name, uploaded_by=uploaded_by, status='PROCESSING')

    wb = load_workbook(file_obj, read_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    raw_headers = _normalize_headers(next(rows, []))
    headers = _map_headers(raw_headers)

    missing = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing:
        logger.error('[IMPORT] ✗ Job %s FAILED — missing columns: %s', job.pk, ', '.join(missing))
        job.status = 'FAILED'
        job.failed_count = 1
        job.save(update_fields=['status', 'failed_count'])
        ImportRowLog.objects.create(
            import_job=job,
            row_number=1,
            status='FAILED',
            error_message=f'Missing columns: {", ".join(missing)}',
        )
        return job

    total_rows = 0
    success_count = 0
    failed_count = 0

    for i, row in enumerate(rows, start=2):
        total_rows += 1
        row_data = _row_to_dict(headers, row)
        email = (row_data.get('email') or '').strip().lower()

        try:
            if not email:
                raise ValueError('Email is required.')

            if User.objects.filter(email=email).exists():
                failed_count += 1
                ImportRowLog.objects.create(
                    import_job=job,
                    row_number=i,
                    email=email,
                    status='SKIPPED',
                    error_message='Duplicate email. User already exists.',
                )
                continue

            username = email.split('@')[0]
            temp_password = generate_temp_password()

            with transaction.atomic():
                sid = transaction.savepoint()
                try:
                    user = User.objects.create_user(
                        username=username,
                        email=email,
                        password=temp_password,
                        is_verified=True,
                        must_change_password=True,
                        parent=uploaded_by,
                    )

                    create_or_update_record(
                        user=user,
                        full_name=(row_data.get('full_name') or '').strip(),
                        phone=(row_data.get('phone') or ''),
                        address=(row_data.get('address') or ''),
                        dob=row_data.get('dob'),
                        source='imported',
                    )
                    transaction.savepoint_commit(sid)
                except Exception:
                    transaction.savepoint_rollback(sid)
                    raise

            send_credentials_email(email, email, temp_password, full_name=(row_data.get('full_name') or '').strip())

            success_count += 1
            logger.info('[IMPORT] Row %d: ✓ %s imported & credentials queued', i, email)
            ImportRowLog.objects.create(
                import_job=job,
                row_number=i,
                email=email,
                status='SUCCESS',
            )
        except Exception as exc:
            failed_count += 1
            logger.error('[IMPORT] Row %d: ✗ %s — %s', i, email or '(no email)', exc)
            ImportRowLog.objects.create(
                import_job=job,
                row_number=i,
                email=email,
                status='FAILED',
                error_message=str(exc),
            )

    job.total_rows = total_rows
    job.success_count = success_count
    job.failed_count = failed_count
    job.status = 'COMPLETED_WITH_ERRORS' if failed_count > 0 else 'COMPLETED'
    job.save(update_fields=['total_rows', 'success_count', 'failed_count', 'status'])
    logger.info('[IMPORT] Job %s %s — %d total, %d success, %d failed', job.pk, job.status, total_rows, success_count, failed_count)
    return job
